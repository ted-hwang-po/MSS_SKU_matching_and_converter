"""기능 B: 날짜별·브랜드별 발주 파일 통합"""
import re
import math
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


def _find_col(df: pd.DataFrame, pattern: str) -> str | None:
    for c in df.columns:
        normalized = re.sub(r'\s+', '', str(c))
        if re.search(pattern, normalized):
            return c
    return None


def _safe_str(val) -> str:
    if val is None:
        return ''
    if isinstance(val, float):
        if math.isnan(val):
            return ''
        if val == int(val):
            return str(int(val))
    return str(val).strip()


def _safe_round(val) -> int | None:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    try:
        return round(float(val))
    except (ValueError, TypeError):
        return None


def _to_int_if_numeric(val):
    if val is None:
        return ''
    if isinstance(val, float):
        if math.isnan(val):
            return ''
        if val == int(val):
            return int(val)
    s = str(val).strip()
    if not s or s == 'nan' or s == 'None':
        return ''
    try:
        f = float(s)
        if f == int(f):
            return int(f)
    except ValueError:
        pass
    return s


def _parse_delivery_date(raw_val) -> str:
    """지정입고일 텍스트를 파싱. '4/1(수)' → '2026-04-01' 등."""
    if raw_val is None:
        return ''
    if isinstance(raw_val, datetime):
        return raw_val.strftime('%Y-%m-%d')

    s = str(raw_val).strip()
    # "4/1(수)" or "3/30(월)" 패턴
    m = re.match(r'(\d{1,2})/(\d{1,2})', s)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        year = datetime.now().year
        try:
            return datetime(year, month, day).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return s


def _extract_delivery_date(filepath: str) -> str:
    """입력 파일에서 지정입고일을 추출 (Row 2, Col B 근처)"""
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    # Row 1~4에서 '지정입고일' 텍스트 탐색
    for row_idx in range(1, 6):
        for col_idx in range(1, 10):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and '지정입고일' in str(val):
                # 바로 옆 셀에 날짜가 있음
                date_val = ws.cell(row=row_idx, column=col_idx + 1).value
                wb.close()
                return _parse_delivery_date(date_val)

    wb.close()
    return ''


def _load_brand_order_file(filepath: str) -> pd.DataFrame:
    """브랜드 송부용 형식의 발주 파일 로드.

    구조: Row 1=빈행, Row 2=지정입고일, Row 3-4=빈행, Row 5=헤더, Row 6=TOTAL, Row 7+=데이터
    Col A=빈, Col B부터 데이터
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    # 헤더 행 찾기: '브랜드명' 또는 '88코드' 가 있는 행
    header_row = None
    for row_idx in range(1, 10):
        for col_idx in range(1, 15):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and '브랜드명' in str(val):
                header_row = row_idx
                break
        if header_row:
            break

    wb.close()

    if header_row is None:
        raise ValueError(f"헤더를 찾을 수 없습니다: {filepath}")

    # pandas로 다시 로드 (헤더 행 지정)
    df = pd.read_excel(filepath, header=header_row - 1, engine='openpyxl')

    # 컬럼명 정규화
    df.columns = [re.sub(r'\s+', ' ', str(c)).strip() for c in df.columns]

    # TOTAL 행 및 빈 행 제거
    if '브랜드명' in df.columns:
        df = df[df['브랜드명'].notna() & (df['브랜드명'] != 'TOTAL')]
    elif _find_col(df, r'브랜드명'):
        brand_col = _find_col(df, r'브랜드명')
        df = df[df[brand_col].notna() & (df[brand_col] != 'TOTAL')]

    # 수량이 0이거나 비어있는 행 제거
    qty_col = _find_col(df, r'발주수량|수량.*오프')
    if qty_col:
        df = df[df[qty_col].notna() & (df[qty_col] != 0)]

    return df


def merge_order_files(
    file_paths: list[str],
    matching_filepath: str | None = None,
    output_path: str | Path | None = None,
) -> pd.DataFrame:
    """여러 발주 파일을 통합.

    Returns: 통합된 DataFrame (출력 파일도 생성)
    """
    all_rows = []

    for fpath in file_paths:
        # 지정입고일 추출
        delivery_date = _extract_delivery_date(fpath)

        # 파일 로드
        df = _load_brand_order_file(fpath)

        # 컬럼 매핑
        brand_col = _find_col(df, r'브랜드명') or '브랜드명'
        barcode_col = _find_col(df, r'88코드|바코드') or '88코드'
        product_no_col = _find_col(df, r'상품번호') or '상품번호'
        name_col = _find_col(df, r'^상품명$|상품명') or '상품명'
        cost_col = _find_col(df, r'공급가.*VAT.*제외') or '공급가(VAT 제외)'
        price_col = _find_col(df, r'정상가.*VAT.*포함') or '정상가(VAT 포함)'
        event_col = _find_col(df, r'상시행사가.*VAT.*포함') or '상시행사가(VAT 포함)'
        qty_col = _find_col(df, r'발주수량|수량.*오프') or '발주수량(오프)'

        for _, row in df.iterrows():
            supply_price = _safe_round(row.get(cost_col, 0)) or 0
            qty = _safe_round(row.get(qty_col, 0)) or 0
            total_amount = supply_price * qty

            all_rows.append({
                '지정입고일': delivery_date,
                '브랜드명': str(row.get(brand_col, '')).strip(),
                '88코드': _to_int_if_numeric(row.get(barcode_col, '')),
                '상품번호': _to_int_if_numeric(row.get(product_no_col, '')),
                '상품명': str(row.get(name_col, '')).strip(),
                '공급가(VAT 제외)': supply_price,
                '정상가(VAT 포함)': _safe_round(row.get(price_col, 0)) or 0,
                '상시행사가(VAT 포함)': _safe_round(row.get(event_col, 0)) or 0,
                '총 발주금액(VAT 제외)': total_amount,
                '발주수량(오프)': qty,
            })

    merged_df = pd.DataFrame(all_rows)

    # 상품코드-바코드 매칭 (선택)
    if matching_filepath:
        merged_df = _fill_product_numbers(merged_df, matching_filepath)

    # 파일 저장
    if output_path:
        _save_merged_output(merged_df, output_path)

    return merged_df


def _fill_product_numbers(merged_df: pd.DataFrame, matching_filepath: str) -> pd.DataFrame:
    """매칭 데이터셋으로 상품번호 채우기."""
    from .loader import load_matching_data

    match_df = load_matching_data(matching_filepath)

    # 바코드 컬럼 찾기
    barcode_col = None
    for c in ['바코드', '88코드']:
        if c in match_df.columns:
            barcode_col = c
            break
    if barcode_col is None:
        return merged_df

    # 상품코드 컬럼 찾기
    uid_col = None
    for c in ['상품코드', '1P 상품코드', '신규 상품코드', '기존 위탁 상품코드 UID']:
        if c in match_df.columns:
            uid_col = c
            break
    if uid_col is None:
        return merged_df

    # 매칭 맵 생성
    barcode_to_uid = {}
    for _, row in match_df.iterrows():
        bc = _safe_str(row[barcode_col])
        uid = _to_int_if_numeric(row[uid_col])
        if bc and uid:
            barcode_to_uid[bc] = uid

    # 상품번호 채우기
    for idx, row in merged_df.iterrows():
        bc = _safe_str(row['88코드'])
        if bc in barcode_to_uid:
            merged_df.at[idx, '상품번호'] = barcode_to_uid[bc]

    return merged_df


def _save_merged_output(df: pd.DataFrame, output_path: str | Path):
    """통합 결과를 샘플 출력 양식에 맞춰 저장."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Row 1~2: 빈 행
    ws.append([])
    ws.append([])

    # Row 3: 헤더 (Col B부터)
    headers = [
        None, '지정입고일', '브랜드명', '88코드', '상품번호', '상품명',
        '공급가(VAT 제외)', '정상가(VAT 포함)', '상시행사가(VAT 포함)',
        '총 발주금액(VAT 제외)', '발주수량(오프)',
    ]
    ws.append(headers)

    # Row 4+: 데이터
    for _, row in df.iterrows():
        # 지정입고일을 datetime으로 변환
        date_val = row.get('지정입고일', '')
        if isinstance(date_val, str) and date_val:
            try:
                date_val = datetime.strptime(date_val, '%Y-%m-%d')
            except ValueError:
                pass

        ws.append([
            None,
            date_val,
            row.get('브랜드명', ''),
            row.get('88코드', ''),
            row.get('상품번호', ''),
            row.get('상품명', ''),
            row.get('공급가(VAT 제외)', 0),
            row.get('정상가(VAT 포함)', 0),
            row.get('상시행사가(VAT 포함)', 0),
            row.get('총 발주금액(VAT 제외)', 0),
            row.get('발주수량(오프)', 0),
        ])

    wb.save(output_path)
