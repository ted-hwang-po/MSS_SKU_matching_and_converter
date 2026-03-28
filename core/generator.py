import math
import re
from copy import copy

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path


def _find_col(df: pd.DataFrame, pattern: str) -> str | None:
    """정규식 패턴으로 컬럼명 검색"""
    for c in df.columns:
        normalized = re.sub(r'\s+', '', str(c))
        if re.search(pattern, normalized):
            return c
    return None


def _to_int_if_numeric(val) -> int | str:
    """숫자 문자열이면 int로 변환, 아니면 그대로 반환"""
    s = str(val).strip()
    if s.replace('.', '', 1).isdigit():
        return int(float(s))
    return s


def _safe_round(val) -> int | None:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    try:
        return round(float(val))
    except (ValueError, TypeError):
        return None


def generate_system_upload(
    merged_df: pd.DataFrame,
    matching_df: pd.DataFrame,
    option_df: pd.DataFrame,
    has_option: dict,
    brand: str,
    output_path: str | Path,
):
    """시스템 업로드용 파일 생성"""
    wb = Workbook()

    # ── 시트 1: 매입상품 구매오더 업로드 양식 ──
    ws1 = wb.active
    ws1.title = '매입상품 구매오더 업로드 양식'

    # 헤더 구조 (샘플 기반)
    # Row 1~6: 빈 행 (메타 정보용)
    # Row 7: 라벨 행 (비제스트 소비가 우선적용, 무신사)
    # Row 8: 필수 마커 행
    # Row 9: 메인 헤더
    # Row 10: 서브 헤더
    # Row 11~: 데이터

    headers_main = [
        '구분', 'UID', '스타일번호', '등급', '이익율', '시작', '종료',
        '오더총 수량', '계획 정상가', '계획 원가',
        'VDC(단가)', 'EPD(단가)', 'MKT(단가)', 'ETC(단가)',
        '납품일', '전자결재', '비고1', '비고2', '비고3',
        '거래처', '선급금 비율', '선급금 지급일', '잔금 지급일',
        '매출 성별', '사이즈유형',
    ]
    # Size01 ~ Size30
    for i in range(1, 31):
        headers_main.append(f'Size{i:02d}')
    headers_main.append('무상품목')

    headers_sub = [
        '조정', '', '', '', '', '판매일', '판매일',
        '(발주)', '(VAT포함)', '(VAT제외)',
        '', '', '', '',
        '', '문서번호', '', '', '',
        '', '', '', '',
        '', 'S06',
    ]
    for i in range(1, 31):
        headers_sub.append(f'{i:02d}')
    headers_sub.append('')

    # 빈 행 (1~6)
    for _ in range(6):
        ws1.append([])

    # Row 7: 라벨
    row7 = [''] * len(headers_main)
    ws1.append(row7)

    # Row 8: 필수 마커
    row8 = [''] * len(headers_main)
    row8[1] = '(필수)'   # UID
    row8[2] = '(필수)'   # 스타일번호
    row8[7] = '(필수)'   # 오더총 수량
    row8[24] = '(필수)'  # 사이즈유형
    ws1.append(row8)

    # Row 9: 메인 헤더
    ws1.append(headers_main)

    # Row 10: 서브 헤더
    ws1.append(headers_sub)

    # 컬럼명 유연 매핑
    qty_col = _find_col(merged_df, r'수량\(?오프\)?') or '수량(오프)'
    price_col = _find_col(merged_df, r'정상가.*VAT.*포함') or '정상가(VAT 포함)'
    cost_col = _find_col(merged_df, r'공급가.*VAT.*제외') or '공급가(VAT 제외)'

    # 데이터: UID 기준으로 그룹핑
    uid_groups = merged_df.groupby('상품코드', sort=False)

    for uid, group in uid_groups:
        uid_str = str(uid)
        first_row = group.iloc[0]
        style_no = str(first_row.get('스타일번호', ''))
        is_opt = has_option.get(uid_str, False)

        # 오더 총 수량 = 동일 UID의 수량(오프) 합산
        total_qty = int(group[qty_col].sum())

        # 가격 (첫 번째 행 기준, 정수 반올림)
        plan_price = _safe_round(first_row.get(price_col, 0))
        plan_cost = _safe_round(first_row.get(cost_col, 0))

        size_type = str(first_row.get('사이즈유형', 'MF3'))
        if not size_type or size_type == 'nan' or size_type == 'None':
            size_type = 'MF3'

        data_row = [''] * len(headers_main)
        data_row[1] = int(float(uid_str)) if uid_str.replace('.', '').isdigit() else uid_str  # UID
        data_row[2] = style_no          # 스타일번호
        data_row[7] = total_qty         # 오더총 수량
        data_row[8] = plan_price        # 계획 정상가
        data_row[9] = plan_cost         # 계획 원가
        data_row[24] = size_type        # 사이즈유형

        # Size 슬롯에 수량 배분
        if is_opt:
            for _, opt_row in group.iterrows():
                slot = opt_row.get('옵션슬롯')
                if slot and not pd.isna(slot):
                    slot_idx = 24 + int(slot)  # Size01 = index 25
                    qty = int(opt_row.get(qty_col, 0))
                    data_row[slot_idx] = qty
        else:
            # 옵션 없는 상품: Size01에 전량
            data_row[25] = total_qty

        ws1.append(data_row)

    # ── 시트 2: Barcode 업로드(옵션) ──
    ws2 = wb.create_sheet('Barcode 업로드(옵션)')
    ws2.append(['스타일번호', 'UID', '사이즈', '바코드'])

    for _, row in merged_df.iterrows():
        uid_str = str(row.get('상품코드', ''))
        style_no_raw = str(row.get('스타일번호', ''))
        barcode_raw = str(row.get('88코드', ''))
        is_opt = has_option.get(uid_str, False)

        if is_opt:
            size_val = str(row.get('옵션값', ''))
            if not size_val or size_val == 'nan' or size_val == 'None':
                size_val = ''
            else:
                # 샘플 양식에 맞춰 옵션값 앞에 공백 추가
                if size_val.startswith('#'):
                    size_val = ' ' + size_val
        else:
            size_val = 'FREE'

        uid_val = _to_int_if_numeric(uid_str)
        style_val = _to_int_if_numeric(style_no_raw)
        barcode_val = _to_int_if_numeric(barcode_raw)

        ws2.append([style_val, uid_val, size_val, barcode_val])

    wb.save(output_path)
    return output_path


def generate_brand_order(
    merged_df: pd.DataFrame,
    has_option: dict,
    brand: str,
    output_path: str | Path,
):
    """브랜드 송부용 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 샘플 양식: A열 비워두고 B열부터 시작
    # Row 1: 빈 행
    ws.append([])

    # Row 2: 지정입고일 (B열)
    ws.append([None, '지정입고일'])

    # Row 3~5: 빈 행
    ws.append([])
    ws.append([])
    ws.append([])

    # Row 6: 헤더 (B열부터)
    headers = [
        None, '브랜드명', '88코드', '상품번호', '상품명',
        '공급가(VAT 제외)', '정상가(VAT 포함)', '상시행사가(VAT 포함)',
        '총 발주금액(VAT 제외)', '발주수량(오프)',
    ]
    ws.append(headers)

    # 컬럼명 유연 매핑
    qty_col = _find_col(merged_df, r'수량\(?오프\)?') or '수량(오프)'
    cost_col = _find_col(merged_df, r'공급가.*VAT.*제외') or '공급가(VAT 제외)'
    price_col = _find_col(merged_df, r'정상가.*VAT.*포함') or '정상가(VAT 포함)'
    event_col = _find_col(merged_df, r'상시행사가.*VAT.*포함') or '상시행사가(VAT 포함)'

    # 데이터 준비
    data_rows = []
    total_amount = 0
    total_qty = 0

    for _, row in merged_df.iterrows():
        uid_str = str(row.get('상품코드', ''))
        barcode = str(row.get('88코드', ''))
        product_name = str(row.get('상품명', ''))

        uid_val = _to_int_if_numeric(uid_str)
        barcode_val = _to_int_if_numeric(barcode)

        supply_price = _safe_round(row.get(cost_col, 0)) or 0
        retail_price = _safe_round(row.get(price_col, 0)) or 0

        event_price_raw = row.get(event_col, 0)
        event_price = _safe_round(event_price_raw) or 0

        qty = int(row.get(qty_col, 0))
        order_amount = supply_price * qty

        total_amount += order_amount
        total_qty += qty

        data_rows.append([
            None, brand, barcode_val, uid_val, product_name,
            supply_price, retail_price, event_price,
            order_amount, qty,
        ])

    # Row 7: TOTAL 행 (B열부터)
    ws.append([
        None, 'TOTAL', None, None, None,
        None, None, None,
        total_amount, total_qty,
    ])

    # Row 8~: 데이터 행
    for dr in data_rows:
        ws.append(dr)

    wb.save(output_path)
    return output_path
