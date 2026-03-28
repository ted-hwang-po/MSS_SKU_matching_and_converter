"""샘플 데이터로 전체 파이프라인 테스트"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from core.loader import load_order_data, load_matching_data, load_option_data, get_brand_list, filter_by_brand
from core.matcher import match_barcode_to_uid, detect_option_products, match_option_info
from core.generator import generate_system_upload, generate_brand_order

SAMPLE_DIR = Path(__file__).parent / 'sample_data'
OUTPUT_DIR = Path(__file__).parent / 'output'
OUTPUT_DIR.mkdir(exist_ok=True)

# 1. 파일 로딩
print("=" * 60)
print("1. 파일 로딩")
print("=" * 60)

order_df = load_order_data(SAMPLE_DIR / '1.총 발주 수량 데이터셋.xlsx')
print(f"  파일1 - 총 발주 수량: {len(order_df)}행")
print(f"  컬럼: {list(order_df.columns)}")

matching_df = load_matching_data(SAMPLE_DIR / '2.상품코드-바코드 매칭 데이터셋.xlsx')
print(f"\n  파일2 - 매칭 데이터: {len(matching_df)}행")
print(f"  컬럼: {list(matching_df.columns)}")

option_df = load_option_data(SAMPLE_DIR / '3.상품별 옵션 정보.xlsx')
print(f"\n  파일3 - 옵션 정보: {len(option_df)}행")
print(f"  컬럼: {list(option_df.columns)}")

# 2. 브랜드 선택
print("\n" + "=" * 60)
print("2. 브랜드 목록")
print("=" * 60)
brands = get_brand_list(order_df)
print(f"  브랜드: {brands}")

brand = brands[0]
print(f"  선택: {brand}")

filtered_df = filter_by_brand(order_df, brand)
print(f"  필터링 후: {len(filtered_df)}행")

# 3. 바코드-UID 매칭
print("\n" + "=" * 60)
print("3. 바코드-UID 매칭")
print("=" * 60)
merged_df, unmatched = match_barcode_to_uid(filtered_df, matching_df)
print(f"  매칭 완료: {len(merged_df)}행")
print(f"  매칭 실패: {unmatched}")
print(f"  매칭 결과 컬럼: {list(merged_df.columns)}")
for _, row in merged_df.iterrows():
    print(f"    88코드={row['88코드']} → UID={row.get('상품코드')}, 스타일번호={row.get('스타일번호')}")

# 4. 옵션 유무 판별
print("\n" + "=" * 60)
print("4. 옵션 유무 판별")
print("=" * 60)
has_option = detect_option_products(merged_df, matching_df)
for uid, is_opt in has_option.items():
    print(f"  UID={uid}: 옵션={'있음' if is_opt else '없음'}")

# 5. 옵션 매핑
print("\n" + "=" * 60)
print("5. 옵션 매핑")
print("=" * 60)
merged_df, warnings = match_option_info(merged_df, option_df, has_option)
for _, row in merged_df.iterrows():
    print(f"  88코드={row['88코드']}, 상품명={row['상품명']}")
    print(f"    → 사이즈유형={row.get('사이즈유형')}, 슬롯={row.get('옵션슬롯')}, 옵션값={row.get('옵션값')}")

if warnings:
    print(f"\n  ⚠️ 경고: {len(warnings)}건")
    for w in warnings:
        print(f"    {w}")
else:
    print("\n  ✅ 경고 없음")

# 6. 파일 생성
print("\n" + "=" * 60)
print("6. 파일 생성")
print("=" * 60)

sys_path = OUTPUT_DIR / f'{brand}_시스템업로드_최종파일.xlsx'
generate_system_upload(merged_df, matching_df, option_df, has_option, brand, sys_path)
print(f"  ✅ 시스템 업로드용: {sys_path}")

brand_path = OUTPUT_DIR / f'{brand}_발주리스트_최종파일.xlsx'
generate_brand_order(merged_df, has_option, brand, brand_path)
print(f"  ✅ 브랜드 송부용: {brand_path}")

# 7. 결과 검증
print("\n" + "=" * 60)
print("7. 결과 검증 (출력 vs 샘플)")
print("=" * 60)

import openpyxl

# 시스템 업로드용 검증
print("\n  [시스템 업로드용]")
wb_out = openpyxl.load_workbook(sys_path, data_only=True)
wb_ref = openpyxl.load_workbook(SAMPLE_DIR / '4.[시스템업로드용] 브랜드명_시스템업로드_최종파일.xlsx', data_only=True)

print(f"  출력 시트: {wb_out.sheetnames}")
print(f"  샘플 시트: {wb_ref.sheetnames}")

# 매입상품 시트 비교
ws_out = wb_out['매입상품 구매오더 업로드 양식']
ws_ref = wb_ref['매입상품 구매오더 업로드 양식']

print(f"\n  [매입상품 구매오더 업로드 양식]")
# 데이터 행 비교 (Row 11~ in ref = Row 9 header + Row 10 subheader)
ref_data_start = 11  # 샘플 파일의 데이터 시작 행
out_data_start = 11  # 출력 파일의 데이터 시작 행

for i in range(3):  # 3개 UID
    ref_row = [ws_ref.cell(row=ref_data_start + i, column=j).value for j in range(1, 60)]
    out_row = [ws_out.cell(row=out_data_start + i, column=j).value for j in range(1, 60)]

    ref_uid = ref_row[1]
    out_uid = out_row[1]
    ref_qty = ref_row[7]
    out_qty = out_row[7]
    ref_size_type = ref_row[24]
    out_size_type = out_row[24]

    match = "✅" if (str(ref_uid) == str(out_uid) and ref_qty == out_qty and ref_size_type == out_size_type) else "❌"
    print(f"  {match} Row {i+1}: UID={out_uid}(ref:{ref_uid}), 수량={out_qty}(ref:{ref_qty}), 사이즈유형={out_size_type}(ref:{ref_size_type})")

    # Size 슬롯 비교
    for s in range(25, 55):
        ref_val = ref_row[s]
        out_val = out_row[s]
        if ref_val or out_val:
            slot_match = "✅" if ref_val == out_val else "❌"
            print(f"      {slot_match} Size{s-24:02d}: {out_val} (ref: {ref_val})")

# Barcode 시트 비교
print(f"\n  [Barcode 업로드(옵션)]")
ws_out2 = wb_out['Barcode 업로드(옵션)']
ws_ref2 = wb_ref['Barcode 업로드(옵션)']

for i in range(2, 6):  # 4 rows
    ref_row = [ws_ref2.cell(row=i, column=j).value for j in range(1, 5)]
    out_row = [ws_out2.cell(row=i, column=j).value for j in range(1, 5)]
    match = "✅" if ref_row == out_row else "❌"
    print(f"  {match} Row {i-1}: 출력={out_row} / 샘플={ref_row}")

# 브랜드 송부용 검증
print(f"\n  [브랜드 송부용]")
wb_out3 = openpyxl.load_workbook(brand_path, data_only=True)
wb_ref3 = openpyxl.load_workbook(SAMPLE_DIR / '5.[브랜드송부용] 브랜드명_발주리스트_최종파일.xlsx', data_only=True)

ws_out3 = wb_out3.active
ws_ref3 = wb_ref3.active

# 데이터 행 비교 (Row 8~ in ref after TOTAL)
for i in range(8, 12):  # 4 data rows
    ref_row = [ws_ref3.cell(row=i, column=j).value for j in range(1, 10)]
    out_row = [ws_out3.cell(row=i, column=j).value for j in range(1, 10)]
    match = "✅" if ref_row == out_row else "❌"
    print(f"  {match} Row {i-7}: 출력={out_row}")
    if match == "❌":
        print(f"         샘플={ref_row}")

wb_out.close()
wb_ref.close()
wb_out3.close()
wb_ref3.close()

print("\n✅ 테스트 완료")
